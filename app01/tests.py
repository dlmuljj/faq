from django.test import TestCase

# Create your tests here.
import os
from openpyxl import Workbook
from openpyxl import load_workbook


def add_data():

    import datetime
    ctime = datetime.datetime.now()

    table_list = [DDR, Serdes, U2, U3, Pll, Hdmi, Dp, TVSensor, MIPI, SARADC, VDAC, Package]
    name_list = ['DDR', 'Serdes', 'U2', 'U3', 'PLL', 'HDMI', 'DP', 'TVSensor', 'MIPI', 'SARADC', 'VDAC', 'Package']

    wb = load_workbook(file_path)
    print(wb.sheetnames)

    for table,name in zip(table_list,name_list):
        print(f'表格{name}导入开始....')
        ws = wb.get_sheet_by_name(name)
        table.objects.all().delete()
        print(f'表格{name}的最大行数为：{ws.max_row}')
        for row in ws.iter_rows(min_row=2,max_col=5, max_row=ws.max_row):
            col_list = []
            for index,cell in enumerate(row):
                # if index == 4:
                #     if cell.value is None:
                #         col_list.append(cell.value)
                #     else:
                #         date_str = cell.value.split('.')
                #         new_value = date_str[0]+'-'+date_str[1]+'-'+date_str[2]
                #         # print(new_value)
                #         col_list.append(new_value)
                # else:
                col_list.append(cell.value)


            if col_list[0] is not None:
                table.objects.create(
                    question=col_list[0],
                    answers=col_list[1],
                    replier=col_list[2],
                    creator=col_list[3],
                    create_time=col_list[4]
                )
            # break
        ddr_l = DDR.objects.all().count()
        print(f'表格{name}的最有效行数为：{ddr_l}')
        print(f'表格{name}导入完毕....\n\n')
        # break

if __name__ == '__main__':
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','faq.settings')
    import django
    django.setup()
    from app01 import models

    module_dir = os.path.dirname(__file__)
    file_path = os.path.join(module_dir,'faq.xlsx')

    from app01.models import DDR, Serdes, U2, U3, Pll, Hdmi, Dp
    from app01.models import TVSensor, MIPI, SARADC, VDAC, Package

    # table_list = [DDR, Serdes, U2, U3, Pll, Hdmi, Dp, TVSensor, MIPI, SARADC, VDAC, Package]
    # for item in table_list:
    # question = models.TextField()
    # answers = models.TextField()
    # replier = models.CharField(max_length=32)
    # creator = models.CharField(max_length=32)
    # create_time = models.DateField(auto_now=False)
    add_data()


