from django.shortcuts import render, redirect, HttpResponse
from django.views import View
from app01 import models
from app01.models import DDR, Serdes, U2, U3, Pll, Hdmi, Dp
from app01.models import TVSensor, MIPI, SARADC, VDAC, Package
# Create your views here.
class Index(View):





    def get(self, request,name=None):

        table_list = [DDR, Serdes, U2, U3, Pll, Hdmi, Dp, TVSensor, MIPI, SARADC, VDAC, Package]
        name_list = ['DDR','Serdes','U2','U3','PLL','HDMI','DP','TVSensor','MIPI','SARADC','VDAC','Package']
        len_list = []
        for item in table_list:
            len_list.append(item.objects.count())
        rst_list = dict(zip(name_list,len_list))

        if name:
            name_upper = name.upper()
        else:
            name_upper = name

        if name_upper == 'DDR':
            ddr_queryset = DDR.objects.all()
        elif name_upper == 'HDMI':
            ddr_queryset = Hdmi.objects.all()
        elif name_upper == 'U2':
            ddr_queryset = U2.objects.all()
        elif name_upper == 'U3':
            ddr_queryset = U3.objects.all()
        elif name_upper =='PLL':
            ddr_queryset = Pll.objects.all()
        elif  name_upper == 'SERDES':
            ddr_queryset = Serdes.objects.all()
        elif name_upper == 'DP':
            ddr_queryset = Dp.objects.all()
        elif name_upper == 'TVSENSOR':
            ddr_queryset = TVSensor.objects.all()
        elif name_upper == 'MIPI':
            ddr_queryset = MIPI.objects.all()
        elif name_upper == 'SARADC':
            ddr_queryset = SARADC.objects.all()
        elif name_upper == 'VDAC':
            ddr_queryset = VDAC.objects.all()
        elif name_upper == 'PACKAGE':
            ddr_queryset = Package.objects.all()
        else:
            ddr_queryset = DDR.objects.all()
        return render(request,'index.html', locals())
        # return HttpResponse('this is a test')



class Ddr(View):
    def get(self, request):
        import os, sys
        from openpyxl import Workbook
        from openpyxl import load_workbook

        # os.chdir(sys.path[0])

        module_dir = os.path.dirname(__file__)
        file_path = os.path.join(module_dir,'faq_1.xlsx')

        wb = Workbook()
        wb = load_workbook(file_path)
        # print(wb.sheetnames)
        ws_ddr = wb.get_sheet_by_name('DDR')
        # print(ws_ddr.max_row)
        # print('ASDF')
        models.DDR.objects.all().delete()
        cnt = 1
        for row in ws_ddr.iter_rows(min_row=2,max_col=5, max_row=32):
            col_list = []
            for cell in row:
                col_list.append(cell.value)
            print('-------'*20)
            print(type(col_list[4]))
            print(col_list[4])
            print('-------'*20)
            models.DDR.objects.create(question=col_list[0],
                                    answers = col_list[1],
                                    replier = col_list[2],
                                    creator = col_list[3],
                                    # create_time = col_list[4],  
            )
            print(f'存储第{cnt}个数据')
            cnt += 1    
        return HttpResponse('done')